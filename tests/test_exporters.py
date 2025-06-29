import pytest
import pandas as pd
import os
from openpyxl import load_workbook
from modules.exporters import export_to_excel
from style_config import COLUMN_WIDTHS, BORDER_CONFIGS, THEMES


class TestExcelExport:
    """Testes para a funcionalidade de exportação Excel"""
    
    @pytest.fixture
    def sample_data(self):
        """Dados de exemplo para testes"""
        return {
            '2024-01': pd.DataFrame({
                'N° OS': ['001', '002'],
                'DATA ENCERRAMENTO': ['2024-01-15', '2024-01-16'],
                'VALOR TOTAL': [1000.50, 2500.75],
                'VALOR MÃO DE OBRA': [500.25, 1200.00],
                'VALOR PEÇAS': [500.25, 1300.75],
                'DESCONTO': [0.00, 100.00],
                'VEÍCULO (PLACA)': ['ABC1234', 'XYZ5678'],
                'CÓDIGO CLIENTE': ['CLI001', 'CLI002'],
                'VALOR PAGO': [1000.50, 2400.75],
                'DEVEDOR': [0.00, 100.00],
                'CARTÃO': [500.25, 1200.00],
                'DINHEIRO': [300.25, 800.75],
                'PIX': [200.00, 400.00],
                'TROCO': [0.00, 0.00],
                'DATA PGTO': ['2024-01-15', '2024-01-16']
            })
        }
    
    @pytest.fixture
    def output_dir(self, tmp_path):
        """Diretório temporário para testes"""
        return str(tmp_path / "test_output")
    
    def test_excel_file_creation(self, sample_data, output_dir):
        """Testa se o arquivo Excel é criado corretamente"""
        export_to_excel(sample_data, output_dir)
        
        # Verifica se o arquivo foi criado
        expected_file = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        assert os.path.exists(expected_file), "Arquivo Excel não foi criado"
    
    def test_worksheet_creation(self, sample_data, output_dir):
        """Testa se a planilha é criada com o nome correto"""
        export_to_excel(sample_data, output_dir)
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        
        # Verifica se a planilha existe
        assert "2024-01" in wb.sheetnames, "Planilha não foi criada com o nome correto"
        
        ws = wb["2024-01"]
        assert ws is not None, "Planilha não foi encontrada"
    
    def test_data_integrity(self, sample_data, output_dir):
        """Testa se os dados foram exportados corretamente"""
        export_to_excel(sample_data, output_dir)
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica se os dados estão corretos
        assert ws['A1'].value == 'N° OS', "Cabeçalho N° OS não encontrado"
        assert ws['A2'].value == '001', "Primeiro valor N° OS incorreto"
        assert ws['A3'].value == '002', "Segundo valor N° OS incorreto"
    
    def test_currency_formatting(self, sample_data, output_dir):
        """Testa se as colunas contábeis recebem formatação de moeda"""
        export_to_excel(sample_data, output_dir)
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica se a coluna VALOR TOTAL tem formatação de moeda
        valor_total_cell = ws['C2']  # VALOR TOTAL, primeira linha de dados
        assert valor_total_cell.number_format == 'R$ #,##0.00', "Formatação de moeda não aplicada"
    
    def test_header_styling(self, sample_data, output_dir):
        """Testa se o cabeçalho recebe a formatação correta"""
        export_to_excel(sample_data, output_dir)
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica estilo do cabeçalho
        header_cell = ws['A1']
        expected_bg = THEMES['default']['header_bg']
        expected_font = THEMES['default']['header_font']
        
        # openpyxl adiciona '00' no início (alpha channel)
        assert header_cell.fill.start_color.rgb[2:] == expected_bg, "Cor de fundo do cabeçalho incorreta"
        assert header_cell.font.color.rgb[2:] == expected_font, "Cor da fonte do cabeçalho incorreta"
        assert header_cell.font.bold, "Cabeçalho não está em negrito"
    
    def test_contabil_styling(self, sample_data, output_dir):
        """Testa se as células contábeis recebem a formatação correta"""
        export_to_excel(sample_data, output_dir)
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica estilo das células contábeis
        contabil_cell = ws['C2']  # VALOR TOTAL, primeira linha de dados
        expected_bg = THEMES['default']['contabil_bg']
        expected_font = THEMES['default']['contabil_font']
        
        # openpyxl adiciona '00' no início (alpha channel)
        assert contabil_cell.fill.start_color.rgb[2:] == expected_bg, "Cor de fundo contábil incorreta"
        assert contabil_cell.font.color.rgb[2:] == expected_font, "Cor da fonte contábil incorreta"
    
    def test_column_widths(self, sample_data, output_dir):
        """Testa se as larguras das colunas são aplicadas corretamente"""
        export_to_excel(sample_data, output_dir)
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica larguras das colunas
        for col_name, expected_width in COLUMN_WIDTHS.items():
            if col_name != 'default':
                # Encontra o índice da coluna
                col_idx = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == col_name:
                        col_idx = idx
                        break
                
                if col_idx:
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    actual_width = ws.column_dimensions[col_letter].width
                    assert actual_width == expected_width, f"Largura da coluna {col_name} incorreta"
    
    def test_border_styling(self, sample_data, output_dir):
        """Testa se as bordas são aplicadas corretamente"""
        export_to_excel(sample_data, output_dir, border_theme='default')
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica bordas do cabeçalho
        header_cell = ws['A1']
        border_config = BORDER_CONFIGS['default']
        
        if border_config['header_border']:
            assert header_cell.border.left.style == border_config['header_border'], "Borda esquerda do cabeçalho incorreta"
            assert header_cell.border.right.style == border_config['header_border'], "Borda direita do cabeçalho incorreta"
            assert header_cell.border.top.style == border_config['header_border'], "Borda superior do cabeçalho incorreta"
            assert header_cell.border.bottom.style == border_config['header_border'], "Borda inferior do cabeçalho incorreta"
    
    def test_border_colors(self, sample_data, output_dir):
        """Testa se as cores das bordas são aplicadas corretamente"""
        export_to_excel(sample_data, output_dir, border_theme='corporate')
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica cor das bordas
        header_cell = ws['A1']
        border_config = BORDER_CONFIGS['corporate']
        
        if border_config['header_border']:
            expected_color = border_config['border_color']
            # openpyxl adiciona '00' no início (alpha channel)
            assert header_cell.border.left.color.rgb[2:] == expected_color, "Cor da borda esquerda incorreta"
            assert header_cell.border.right.color.rgb[2:] == expected_color, "Cor da borda direita incorreta"
            assert header_cell.border.top.color.rgb[2:] == expected_color, "Cor da borda superior incorreta"
            assert header_cell.border.bottom.color.rgb[2:] == expected_color, "Cor da borda inferior incorreta"
    
    def test_multiple_border_themes(self, sample_data, output_dir):
        """Testa diferentes temas de bordas"""
        themes_to_test = ['default', 'corporate', 'dark', 'minimal']
        
        for theme in themes_to_test:
            # Limpa diretório para cada teste
            if os.path.exists(output_dir):
                for file in os.listdir(output_dir):
                    os.remove(os.path.join(output_dir, file))
            
            export_to_excel(sample_data, output_dir, border_theme=theme)
            
            file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
            assert os.path.exists(file_path), f"Arquivo não criado para tema {theme}"
            
            wb = load_workbook(file_path)
            ws = wb["2024-01"]
            
            # Verifica se pelo menos o cabeçalho tem bordas (exceto minimal)
            header_cell = ws['A1']
            if theme != 'minimal':
                assert header_cell.border.left.style is not None, f"Bordas não aplicadas para tema {theme}"
    
    def test_currency_formats(self, sample_data, output_dir):
        """Testa diferentes formatos de moeda"""
        currencies = ['BRL', 'USD', 'EUR']
        
        for currency in currencies:
            # Limpa diretório para cada teste
            if os.path.exists(output_dir):
                for file in os.listdir(output_dir):
                    os.remove(os.path.join(output_dir, file))
            
            export_to_excel(sample_data, output_dir, currency=currency)
            
            file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
            wb = load_workbook(file_path)
            ws = wb["2024-01"]
            
            # Verifica formato de moeda
            valor_cell = ws['C2']  # VALOR TOTAL
            if currency == 'BRL':
                assert valor_cell.number_format == 'R$ #,##0.00', "Formato BRL incorreto"
            elif currency == 'USD':
                assert valor_cell.number_format == 'US$ #,##0.00', "Formato USD incorreto"
            elif currency == 'EUR':
                assert valor_cell.number_format == '€ #,##0.00', "Formato EUR incorreto"
    
    def test_theme_colors(self, sample_data, output_dir):
        """Testa diferentes temas de cores"""
        themes_to_test = ['default', 'dark']
        
        for theme in themes_to_test:
            # Limpa diretório para cada teste
            if os.path.exists(output_dir):
                for file in os.listdir(output_dir):
                    os.remove(os.path.join(output_dir, file))
            
            export_to_excel(sample_data, output_dir, theme=theme)
            
            file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
            wb = load_workbook(file_path)
            ws = wb["2024-01"]
            
            # Verifica cores do tema
            header_cell = ws['A1']
            theme_config = THEMES[theme]
            
            # openpyxl adiciona '00' no início (alpha channel)
            assert header_cell.fill.start_color.rgb[2:] == theme_config['header_bg'], f"Cor de fundo do cabeçalho incorreta para tema {theme}"
            assert header_cell.font.color.rgb[2:] == theme_config['header_font'], f"Cor da fonte do cabeçalho incorreta para tema {theme}"
    
    def test_empty_dataframe(self, output_dir):
        """Testa exportação com DataFrame vazio"""
        empty_data = {'2024-01': pd.DataFrame()}
        
        export_to_excel(empty_data, output_dir)
        
        file_path = os.path.join(output_dir, "Recebimentos_2024-01.xlsx")
        assert os.path.exists(file_path), "Arquivo não criado para DataFrame vazio"
        
        wb = load_workbook(file_path)
        ws = wb["2024-01"]
        
        # Verifica se a planilha existe mesmo vazia
        assert ws is not None, "Planilha não criada para DataFrame vazio"
    
    def test_invalid_output_dir(self, sample_data):
        """Testa comportamento com diretório de saída inválido"""
        invalid_dir = "/caminho/inexistente/para/teste"
        
        # Deve criar o diretório automaticamente
        export_to_excel(sample_data, invalid_dir)
        
        # Verifica se o diretório foi criado
        assert os.path.exists(invalid_dir), "Diretório não foi criado automaticamente"
        
        # Limpa após o teste
        import shutil
        shutil.rmtree(invalid_dir, ignore_errors=True) 