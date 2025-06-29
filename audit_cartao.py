#!/usr/bin/env python3
"""
Auditoria Especializada para Transações de Cartão
Compara dados do CSV de extratos com dados gerados pela aplicação
"""

import os
import logging
import pandas as pd
from datetime import datetime
from modules.auditor import DataAuditor, AuditError
from config import OUTPUT_DIR
from style_config import (
    COLUMN_WIDTHS, BORDER_CONFIGS, THEMES, 
    CURRENCY_FORMATS, DATE_FORMATS, CONTABEIS_COLS
)


def setup_logging():
    """Configura logging para a auditoria de cartão"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)


def parse_cartao_csv(csv_file_path: str) -> pd.DataFrame:
    """
    Carrega e processa o CSV de transações de cartão
    
    Args:
        csv_file_path: Caminho para o arquivo CSV
        
    Returns:
        pd.DataFrame: Dados processados do CSV
    """
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Carregando CSV de transações: {csv_file_path}")
        
        # Carrega o CSV
        df = pd.read_csv(csv_file_path, encoding='utf-8')
        
        # Processa a coluna de data
        df['Data e hora'] = pd.to_datetime(df['Data e hora'], format='%d %b, %Y · %H:%M')
        df['DATA_PGTO'] = df['Data e hora'].dt.date
        
        # Processa valores monetários (remove aspas e converte para float)
        df['Valor (R$)'] = df['Valor (R$)'].str.replace('"', '').str.replace('.', '').str.replace(',', '.').astype(float)
        df['Líquido (R$)'] = df['Líquido (R$)'].str.replace('"', '').str.replace('.', '').str.replace(',', '.').astype(float)
        
        # Cria colunas para auditoria
        df['TIPO_PAGAMENTO'] = df['Meio - Meio'].apply(lambda x: 'CARTÃO' if x in ['Crédito', 'Débito', 'Credito', 'Debito'] else 'PIX')
        df['VALOR_AUDITORIA'] = df['Valor (R$)']
        
        logger.info(f"CSV processado: {len(df)} transações")
        logger.info(f"Transações por tipo: {df['TIPO_PAGAMENTO'].value_counts().to_dict()}")
        
        return df
        
    except Exception as e:
        error_msg = f"Erro ao processar CSV de transações: {e}"
        logger.error(error_msg)
        raise AuditError(error_msg)


def create_audit_mappings(cartao_df: pd.DataFrame) -> dict:
    """
    Cria mapeamentos específicos para auditoria baseado no tipo de pagamento
    
    Args:
        cartao_df: DataFrame com transações de cartão
        
    Returns:
        dict: Mapeamentos para auditoria
    """
    mappings = {}
    
    # Para cada transação, cria um mapeamento específico
    for idx, row in cartao_df.iterrows():
        identificador = row['Identificador']
        tipo_pagamento = row['TIPO_PAGAMENTO']
        valor = row['VALOR_AUDITORIA']
        data_pgto = row['DATA_PGTO']
        
        if tipo_pagamento == 'CARTÃO':
            # Mapeia para coluna CARTÃO
            mappings[identificador] = {
                'csv_field': 'VALOR_AUDITORIA',
                'generated_field': 'CARTÃO',
                'valor': valor,
                'data_pgto': data_pgto,
                'tipo': 'CARTÃO'
            }
        else:
            # Mapeia para coluna PIX
            mappings[identificador] = {
                'csv_field': 'VALOR_AUDITORIA',
                'generated_field': 'PIX',
                'valor': valor,
                'data_pgto': data_pgto,
                'tipo': 'PIX'
            }
    
    return mappings


def audit_cartao_transactions(csv_file_path: str, generated_file_path: str, output_report: str = None):
    """
    Executa auditoria especializada para transações de cartão
    
    Args:
        csv_file_path: Caminho para o CSV de transações
        generated_file_path: Caminho para o arquivo Excel gerado
        output_report: Caminho para o relatório de saída
    """
    logger = setup_logging()
    
    try:
        logger.info("=== AUDITORIA DE TRANSAÇÕES DE CARTÃO ===")
        
        # Verifica se os arquivos existem
        if not os.path.exists(csv_file_path):
            logger.error(f"Arquivo CSV não encontrado: {csv_file_path}")
            return
        
        if not os.path.exists(generated_file_path):
            logger.error(f"Arquivo gerado não encontrado: {generated_file_path}")
            return
        
        # Carrega dados
        cartao_df = parse_cartao_csv(csv_file_path)
        auditor = DataAuditor(tolerance_percentage=0.01)  # 1% de tolerância
        
        # Carrega dados gerados
        generated_df = auditor.load_generated_data(generated_file_path)
        generated_df = auditor.normalize_column_names(generated_df)
        
        # Converte DATA PGTO para date se necessário
        if 'DATA PGTO' in generated_df.columns:
            generated_df['DATA PGTO'] = pd.to_datetime(generated_df['DATA PGTO']).dt.date
        
        logger.info(f"Dados gerados carregados: {len(generated_df)} registros")
        
        # Executa auditoria
        results = []
        summary_stats = {
            'total_transacoes': len(cartao_df),
            'cartao_encontradas': 0,
            'pix_encontradas': 0,
            'nao_encontradas': 0,
            'valores_coincidentes': 0,
            'valores_divergentes': 0
        }
        
        for idx, cartao_row in cartao_df.iterrows():
            identificador = cartao_row['Identificador']
            valor_cartao = cartao_row['VALOR_AUDITORIA']
            data_cartao = cartao_row['DATA_PGTO']
            tipo_pagamento = cartao_row['TIPO_PAGAMENTO']
            
            # Procura registro correspondente por data
            matching_generated = generated_df[generated_df['DATA PGTO'] == data_cartao]
            
            if len(matching_generated) == 0:
                # Transação não encontrada
                results.append({
                    'identificador': identificador,
                    'data_cartao': data_cartao,
                    'valor_cartao': valor_cartao,
                    'tipo_pagamento': tipo_pagamento,
                    'status': 'NÃO ENCONTRADA',
                    'valor_gerado': None,
                    'diferenca': None,
                    'observacao': f'Data {data_cartao} não encontrada nos dados gerados'
                })
                summary_stats['nao_encontradas'] += 1
                continue
            
            # Procura por valor na coluna correspondente
            campo_procurado = 'CARTÃO' if tipo_pagamento == 'CARTÃO' else 'PIX'
            valor_encontrado = None
            
            for _, gen_row in matching_generated.iterrows():
                if campo_procurado in gen_row.index:
                    valor_gen = gen_row[campo_procurado]
                    if pd.notna(valor_gen) and abs(valor_gen - valor_cartao) <= (valor_cartao * 0.01):  # 1% tolerância
                        valor_encontrado = valor_gen
                        break
            
            if valor_encontrado is not None:
                # Valor encontrado
                diferenca = abs(valor_cartao - valor_encontrado)
                is_match = diferenca <= (valor_cartao * 0.01)
                
                results.append({
                    'identificador': identificador,
                    'data_cartao': data_cartao,
                    'valor_cartao': valor_cartao,
                    'tipo_pagamento': tipo_pagamento,
                    'status': 'COINCIDENTE' if is_match else 'DIVERGENTE',
                    'valor_gerado': valor_encontrado,
                    'diferenca': diferenca,
                    'observacao': f'Encontrado na coluna {campo_procurado}'
                })
                
                if tipo_pagamento == 'CARTÃO':
                    summary_stats['cartao_encontradas'] += 1
                else:
                    summary_stats['pix_encontradas'] += 1
                
                if is_match:
                    summary_stats['valores_coincidentes'] += 1
                else:
                    summary_stats['valores_divergentes'] += 1
            else:
                # Valor não encontrado
                results.append({
                    'identificador': identificador,
                    'data_cartao': data_cartao,
                    'valor_cartao': valor_cartao,
                    'tipo_pagamento': tipo_pagamento,
                    'status': 'VALOR NÃO ENCONTRADO',
                    'valor_gerado': None,
                    'diferenca': None,
                    'observacao': f'Valor {valor_cartao} não encontrado na coluna {campo_procurado} para a data {data_cartao}'
                })
                summary_stats['nao_encontradas'] += 1
        
        # Exibe resumo
        logger.info("\n=== RESUMO DA AUDITORIA ===")
        logger.info(f"Total de transações: {summary_stats['total_transacoes']}")
        logger.info(f"Cartão encontradas: {summary_stats['cartao_encontradas']}")
        logger.info(f"PIX encontradas: {summary_stats['pix_encontradas']}")
        logger.info(f"Não encontradas: {summary_stats['nao_encontradas']}")
        logger.info(f"Valores coincidentes: {summary_stats['valores_coincidentes']}")
        logger.info(f"Valores divergentes: {summary_stats['valores_divergentes']}")
        
        taxa_sucesso = (summary_stats['valores_coincidentes'] / summary_stats['total_transacoes']) * 100 if summary_stats['total_transacoes'] > 0 else 0
        logger.info(f"Taxa de sucesso: {taxa_sucesso:.2f}%")
        
        # Exibe algumas divergências
        divergencias = [r for r in results if r['status'] in ['DIVERGENTE', 'NÃO ENCONTRADA', 'VALOR NÃO ENCONTRADO']]
        if divergencias:
            logger.info(f"\n=== PRIMEIRAS 5 DIVERGÊNCIAS ===")
            for i, result in enumerate(divergencias[:5]):
                logger.info(f"{i+1}. ID: {result['identificador']}")
                logger.info(f"   Data: {result['data_cartao']}")
                logger.info(f"   Tipo: {result['tipo_pagamento']}")
                logger.info(f"   Valor CSV: R$ {result['valor_cartao']:.2f}")
                logger.info(f"   Valor Gerado: R$ {result['valor_gerado']:.2f}" if result['valor_gerado'] else "   Valor Gerado: N/A")
                logger.info(f"   Status: {result['status']}")
                logger.info(f"   Observação: {result['observacao']}")
                logger.info("")
        
        # Gera relatório Excel se solicitado
        if output_report:
            generate_cartao_report(results, summary_stats, output_report)
            logger.info(f"📊 Relatório salvo em: {output_report}")
        
        logger.info("✅ Auditoria de cartão concluída!")
        
    except Exception as e:
        logger.error(f"❌ Erro na auditoria: {e}")
        raise


def generate_cartao_report(results: list, summary_stats: dict, output_file: str):
    """
    Gera relatório Excel da auditoria de cartão, detalhando divergências e problemas
    Args:
        results: Lista de resultados da auditoria
        summary_stats: Estatísticas resumidas
        output_file: Caminho para o arquivo de saída
    """
    try:
        # Garante que a pasta existe
        pasta = os.path.dirname(output_file)
        if pasta and not os.path.exists(pasta):
            os.makedirs(pasta)

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Configurações de estilo
            theme = THEMES['default']
            border_config = BORDER_CONFIGS['default']
            
            # Resumo
            summary_data = {
                'Métrica': [
                    'Total de Transações',
                    'Cartão Encontradas',
                    'PIX Encontradas',
                    'Não Encontradas',
                    'Valores Coincidentes',
                    'Valores Divergentes',
                    'Taxa de Sucesso (%)',
                    'Data da Auditoria'
                ],
                'Valor': [
                    summary_stats['total_transacoes'],
                    summary_stats['cartao_encontradas'],
                    summary_stats['pix_encontradas'],
                    summary_stats['nao_encontradas'],
                    summary_stats['valores_coincidentes'],
                    summary_stats['valores_divergentes'],
                    f"{(summary_stats['valores_coincidentes'] / summary_stats['total_transacoes']) * 100:.2f}%" if summary_stats['total_transacoes'] > 0 else "0%",
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Resumo', index=False)
            
            # Aplica formatação ao resumo
            worksheet = writer.sheets['Resumo']
            apply_worksheet_formatting(worksheet, summary_df, theme, border_config)

            # Detalhes (todas as transações)
            details_df = pd.DataFrame(results)
            # Adiciona coluna de diferença percentual
            details_df['dif_percentual'] = details_df.apply(
                lambda row: (row['diferenca'] / row['valor_cartao'] * 100) if row['diferenca'] is not None and row['valor_cartao'] else None, axis=1)
            # Reorganiza colunas para facilitar análise
            colunas = [
                'identificador', 'data_cartao', 'tipo_pagamento', 'valor_cartao', 'valor_gerado',
                'diferenca', 'dif_percentual', 'status', 'observacao'
            ]
            details_df = details_df[[c for c in colunas if c in details_df.columns]]
            details_df.to_excel(writer, sheet_name='Detalhes', index=False)
            
            # Aplica formatação aos detalhes
            worksheet = writer.sheets['Detalhes']
            apply_worksheet_formatting(worksheet, details_df, theme, border_config)

            # Divergências detalhadas
            divergencias = [r for r in results if r['status'] in ['DIVERGENTE', 'NÃO ENCONTRADA', 'VALOR NÃO ENCONTRADO']]
            if divergencias:
                divergencias_df = pd.DataFrame(divergencias)
                divergencias_df['dif_percentual'] = divergencias_df.apply(
                    lambda row: (row['diferenca'] / row['valor_cartao'] * 100) if row['diferenca'] is not None and row['valor_cartao'] else None, axis=1)
                divergencias_df = divergencias_df[[c for c in colunas if c in divergencias_df.columns]]
                divergencias_df.to_excel(writer, sheet_name='Divergências', index=False)
                
                # Aplica formatação às divergências
                worksheet = writer.sheets['Divergências']
                apply_worksheet_formatting(worksheet, divergencias_df, theme, border_config)

            # Nova aba: Problemas detalhados
            problemas = []
            for r in results:
                if r['status'] in ['DIVERGENTE', 'NÃO ENCONTRADA', 'VALOR NÃO ENCONTRADO']:
                    problemas.append({
                        'Identificador': r['identificador'],
                        'Data': r['data_cartao'],
                        'Tipo': r['tipo_pagamento'],
                        'Valor CSV': r['valor_cartao'],
                        'Valor Gerado': r['valor_gerado'],
                        'Diferença Absoluta': r['diferenca'],
                        'Diferença Percentual': (r['diferenca'] / r['valor_cartao'] * 100) if r['diferenca'] is not None and r['valor_cartao'] else None,
                        'Status': r['status'],
                        'Observação': r['observacao']
                    })
            if problemas:
                problemas_df = pd.DataFrame(problemas)
                problemas_df.to_excel(writer, sheet_name='Problemas Detalhados', index=False)
                
                # Aplica formatação aos problemas detalhados
                worksheet = writer.sheets['Problemas Detalhados']
                apply_worksheet_formatting(worksheet, problemas_df, theme, border_config)

    except Exception as e:
        logging.error(f"Erro ao gerar relatório: {e}")
        raise


def apply_worksheet_formatting(worksheet, df, theme, border_config):
    """
    Aplica formatação uniforme à planilha baseada no style_config
    
    Args:
        worksheet: Planilha do openpyxl
        df: DataFrame com os dados
        theme: Tema de cores
        border_config: Configuração de bordas
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    # Configurações de borda
    border_style = Side(style=border_config['data_border'], color=border_config['border_color'])
    header_border_style = Side(style=border_config['header_border'], color=border_config['border_color'])
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color=theme['header_font'])
    header_fill = PatternFill(start_color=theme['header_bg'], end_color=theme['header_bg'], fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Estilo das células de dados
    data_font = Font(color='000000')
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    # Aplica formatação ao cabeçalho
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(
            left=header_border_style, right=header_border_style,
            top=header_border_style, bottom=header_border_style
        )
    
    # Aplica formatação aos dados
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = Border(
                left=border_style, right=border_style,
                top=border_style, bottom=border_style
            )
            
            # Formatação específica para colunas numéricas
            column_name = df.columns[col - 1]
            if any(keyword in column_name.lower() for keyword in ['valor', 'diferenca', 'percentual']):
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = CURRENCY_FORMATS['BRL']
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Formatação para datas
            elif 'data' in column_name.lower():
                if cell.value is not None:
                    cell.number_format = DATE_FORMATS['pt_BR']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajusta largura das colunas
    for col in range(1, len(df.columns) + 1):
        column_name = df.columns[col - 1]
        width = COLUMN_WIDTHS.get(column_name, COLUMN_WIDTHS['default'])
        worksheet.column_dimensions[get_column_letter(col)].width = width


def main():
    """Função principal"""
    logger = setup_logging()
    
    try:
        # Configurações
        csv_file = "data/extratos/report_20250628_194465.csv"
        generated_file = os.path.join(OUTPUT_DIR, "Recebimentos_2025-06.xlsx")  # Ajuste conforme necessário
        report_file = "data/relatorios/auditoria_cartao_relatorio.xlsx"
        
        logger.info("Iniciando auditoria de transações de cartão...")
        logger.info(f"CSV: {csv_file}")
        logger.info(f"Gerado: {generated_file}")
        
        # Executa auditoria
        audit_cartao_transactions(csv_file, generated_file, report_file)
        
    except Exception as e:
        logger.error(f"Erro inesperado: {e}")


if __name__ == '__main__':
    main() 