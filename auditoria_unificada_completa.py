#!/usr/bin/env python3
"""
Auditoria Unificada Completa - Cartão e PIX
Combina auditoria de transações de cartão e PIX em um único relatório Excel
"""

import os
import logging
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from modules.auditor import DataAuditor, AuditError
from config import OUTPUT_DIR
from style_config import (
    COLUMN_WIDTHS, BORDER_CONFIGS, THEMES, 
    CURRENCY_FORMATS, DATE_FORMATS, CONTABEIS_COLS
)
import json
import unicodedata


@dataclass
class PixTransaction:
    """Representa uma transação PIX"""
    data: str
    valor: float
    descricao: str
    origem: str  # 'banco', 'cartao', 'recebimentos'
    identificador: Optional[str] = None
    referencia: Optional[str] = None
    remetente: Optional[str] = None  # Nome ou CPF do remetente
    chave_pix: Optional[str] = None  # Chave PIX do remetente


@dataclass
class GroupedPixTransaction:
    """Representa transações PIX agrupadas por remetente e data"""
    data: str
    valor_total: float
    remetente: str
    origem: str
    transacoes_originais: List[PixTransaction]
    quantidade_transacoes: int
    referencia: Optional[str] = None


@dataclass
class AuditMatch:
    """Resultado de uma correspondência encontrada"""
    banco_transaction: PixTransaction
    recebimentos_transaction: Optional[PixTransaction] = None
    cartao_transaction: Optional[PixTransaction] = None
    match_type: str = "exato"
    confidence: float = 1.0
    notes: str = ""


def setup_logging():
    """Configura logging para a auditoria unificada"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)


def select_files_gui():
    """Interface gráfica para seleção de arquivos"""
    # Cria janela principal (oculta)
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal
    
    files = {}
    
    try:
        # Seleciona arquivo CSV do cartão
        messagebox.showinfo("Seleção de Arquivos", 
                          "Selecione o arquivo CSV de transações do cartão")
        cartao_csv = filedialog.askopenfilename(
            title="Selecione o arquivo CSV do cartão",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if not cartao_csv:
            messagebox.showerror("Erro", "Nenhum arquivo CSV do cartão selecionado!")
            return None
        
        files['cartao_csv'] = cartao_csv
        
        # Seleciona arquivo CSV do banco
        messagebox.showinfo("Seleção de Arquivos", 
                          "Selecione o arquivo CSV de transações PIX do banco")
        banco_csv = filedialog.askopenfilename(
            title="Selecione o arquivo CSV do banco",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if not banco_csv:
            messagebox.showerror("Erro", "Nenhum arquivo CSV do banco selecionado!")
            return None
        
        files['banco_csv'] = banco_csv
        
        # Seleciona arquivo Excel de recebimentos
        messagebox.showinfo("Seleção de Arquivos", 
                          "Selecione o arquivo Excel de recebimentos")
        recebimentos_excel = filedialog.askopenfilename(
            title="Selecione o arquivo Excel de recebimentos",
            filetypes=[("Excel files", "*.xlsx"), ("Excel files", "*.xls"), ("All files", "*.*")]
        )
        
        if not recebimentos_excel:
            messagebox.showerror("Erro", "Nenhum arquivo Excel de recebimentos selecionado!")
            return None
        
        files['recebimentos_excel'] = recebimentos_excel
        
        # Confirma seleção
        confirm_msg = f"""
Arquivos selecionados:

📄 Cartão: {os.path.basename(cartao_csv)}
🏦 Banco: {os.path.basename(banco_csv)}
📊 Recebimentos: {os.path.basename(recebimentos_excel)}

Deseja continuar com a auditoria?
        """
        
        if messagebox.askyesno("Confirmar Arquivos", confirm_msg):
            return files
        else:
            messagebox.showinfo("Cancelado", "Auditoria cancelada pelo usuário")
            return None
            
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao selecionar arquivos: {e}")
        return None
    finally:
        root.destroy()


def select_files_powershell():
    """Seleção de arquivos via PowerShell (fallback)"""
    logger = logging.getLogger(__name__)
    
    print("\n=== SELEÇÃO DE ARQUIVOS ===")
    print("Digite os caminhos dos arquivos ou pressione Enter para usar os padrões:")
    
    # Arquivo CSV do cartão
    cartao_csv = input(f"CSV do cartão (padrão: data/extratos/report_20250628_194465.csv): ").strip()
    if not cartao_csv:
        cartao_csv = "data/extratos/report_20250628_194465.csv"
    
    # Arquivo CSV do banco
    banco_csv = input(f"CSV do banco (padrão: data/extratos/NU_636868111_01JUN2025_27JUN2025.csv): ").strip()
    if not banco_csv:
        banco_csv = "data/extratos/NU_636868111_01JUN2025_27JUN2025.csv"
    
    # Arquivo Excel de recebimentos
    recebimentos_excel = input(f"Excel de recebimentos (padrão: data/recebimentos/Recebimentos_2025-06.xlsx): ").strip()
    if not recebimentos_excel:
        recebimentos_excel = "data/recebimentos/Recebimentos_2025-06.xlsx"
    
    return {
        'cartao_csv': cartao_csv,
        'banco_csv': banco_csv,
        'recebimentos_excel': recebimentos_excel
    }


def parse_cartao_csv(csv_file_path: str) -> pd.DataFrame:
    """
    Carrega e processa o CSV de transações de cartão
    """
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Carregando CSV de transações: {csv_file_path}")
        
        # Carrega o CSV
        df = pd.read_csv(csv_file_path, encoding='utf-8')
        
        # Processa a coluna de data
        df['Data e hora'] = pd.to_datetime(df['Data e hora'], format='%d %b, %Y · %H:%M')
        df['DATA_PGTO'] = df['Data e hora'].dt.date
        
        # Processa valores monetários
        df['Valor (R$)'] = df['Valor (R$)'].str.replace('"', '').str.replace('.', '').str.replace(',', '.').astype(float)
        df['Líquido (R$)'] = df['Líquido (R$)'].str.replace('"', '').str.replace('.', '').str.replace(',', '.').astype(float)
        
        # Cria colunas para auditoria
        df['TIPO_PAGAMENTO'] = df['Meio - Meio'].apply(lambda x: 'CARTÃO' if x in ['Crédito', 'Débito', 'Credito', 'Debito'] else 'PIX')
        df['VALOR_AUDITORIA'] = df['Valor (R$)']
        
        logger.info(f"CSV processado: {len(df)} transações")
        logger.info(f"Transações por tipo: {df['TIPO_PAGAMENTO'].value_counts().to_dict()}")
        
        return df
        
    except Exception as e:
        error_msg = f"Erro ao processar CSV de transações: {e}"
        logger.error(error_msg)
        raise AuditError(error_msg)


def load_banco_pix_json(json_path: str) -> List[PixTransaction]:
    """Carrega transações PIX do JSON do banco (já filtrado na conversão)"""
    logger = logging.getLogger(__name__)
    logger.info(f"Carregando JSON do banco: {json_path}")
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        transactions = []
        
        for row in data:
            try:
                valor = float(str(row.get('Valor', '0')).replace(',', '.'))
                data_tx = str(row.get('Data', '')).strip()
                descricao = str(row.get('Descrição', '')).strip()
                
                # Normaliza a data para formato DD/MM/YYYY para comparação
                try:
                    data_dt = pd.to_datetime(data_tx, dayfirst=True)  # DD/MM/YYYY
                    data_normalizada = data_dt.strftime('%d/%m/%Y')
                except Exception:
                    data_normalizada = data_tx
                    
                remetente = extract_remetente_from_description(descricao)
                chave_pix = extract_chave_pix_from_description(descricao)
                
                transaction = PixTransaction(
                    data=data_normalizada,
                    valor=valor,
                    descricao=descricao,
                    origem='banco',
                    identificador=None,
                    remetente=remetente,
                    chave_pix=chave_pix
                )
                transactions.append(transaction)
                
            except (ValueError, KeyError) as e:
                logger.warning(f"Erro ao processar linha do banco: {e}")
                continue
        logger.info(f"Carregadas {len(transactions)} transações PIX do banco JSON")
        return transactions
        
    except Exception as e:
        logger.error(f"Erro ao carregar JSON do banco: {e}")
        return []


def load_recebimentos_json(json_path: str) -> List[PixTransaction]:
    """Carrega transações PIX da tabela de recebimentos em formato JSON"""
    logger = logging.getLogger(__name__)
    logger.info(f"Carregando JSON de recebimentos: {json_path}")
    
    try:
        df = pd.read_json(json_path, orient='records')
        transactions = []
        
        for _, row in df.iterrows():
            try:
                # Verifica se tem valor PIX
                valor_pix = row.get('PIX', 0)
                if pd.notna(valor_pix) and float(valor_pix) > 0:
                    data_pgto = str(row.get('DATA PGTO', '')).strip()
                    if data_pgto and data_pgto != 'nan':
                        # Normaliza a data para formato DD/MM/YYYY para comparação
                        try:
                            # Converte para datetime e depois para string no formato desejado
                            if 'T' in data_pgto:  # Formato ISO
                                data_dt = pd.to_datetime(data_pgto)
                            else:  # Outros formatos
                                data_dt = pd.to_datetime(data_pgto, format='mixed')
                            data_normalizada = data_dt.strftime('%d/%m/%Y')
                        except:
                            # Se falhar a conversão, mantém o formato original
                            data_normalizada = data_pgto
                        
                        transaction = PixTransaction(
                            data=data_normalizada,
                            valor=float(valor_pix),
                            descricao=f"Recebimento PIX - OS: {row.get('N° OS', 'N/A')}",
                            origem='recebimentos',
                            referencia=str(row.get('N° OS', '')).strip()
                        )
                        transactions.append(transaction)
                        
            except (ValueError, KeyError) as e:
                logger.warning(f"Erro ao processar linha de recebimentos: {e}")
                continue
        
        logger.info(f"Carregadas {len(transactions)} transações PIX dos recebimentos")
        return transactions
        
    except Exception as e:
        logger.error(f"Erro ao carregar JSON de recebimentos: {e}")
        return []


def load_recebimentos_excel(excel_path: str) -> List[PixTransaction]:
    """Carrega transações PIX da tabela de recebimentos (mantido para compatibilidade)"""
    logger = logging.getLogger(__name__)
    logger.info(f"Carregando Excel de recebimentos: {excel_path}")
    
    try:
        df = pd.read_excel(excel_path)
        transactions = []
        
        for _, row in df.iterrows():
            try:
                # Verifica se tem valor PIX
                valor_pix = row.get('PIX', 0)
                if pd.notna(valor_pix) and float(valor_pix) > 0:
                    data_pgto = str(row.get('DATA PGTO', '')).strip()
                    if data_pgto and data_pgto != 'nan':
                        transaction = PixTransaction(
                            data=data_pgto,
                            valor=float(valor_pix),
                            descricao=f"Recebimento PIX - OS: {row.get('N° OS', 'N/A')}",
                            origem='recebimentos',
                            referencia=str(row.get('N° OS', '')).strip()
                        )
                        transactions.append(transaction)
                        
            except (ValueError, KeyError) as e:
                logger.warning(f"Erro ao processar linha de recebimentos: {e}")
                continue
        
        logger.info(f"Carregadas {len(transactions)} transações PIX dos recebimentos")
        return transactions
        
    except Exception as e:
        logger.error(f"Erro ao carregar Excel de recebimentos: {e}")
        return []


def audit_cartao_transactions(cartao_df: pd.DataFrame, generated_df: pd.DataFrame) -> List[Dict]:
    """Executa auditoria de transações de cartão"""
    results = []
    
    for idx, cartao_row in cartao_df.iterrows():
        identificador = cartao_row['Identificador']
        valor_cartao = cartao_row['VALOR_AUDITORIA']
        data_cartao = cartao_row['DATA_PGTO']
        tipo_pagamento = cartao_row['TIPO_PAGAMENTO']
        
        # Procura registro correspondente por data
        matching_generated = generated_df[generated_df['DATA PGTO'] == data_cartao]
        
        if len(matching_generated) == 0:
            # Transação não encontrada
            results.append({
                'identificador': identificador,
                'data_cartao': data_cartao,
                'valor_cartao': valor_cartao,
                'tipo_pagamento': tipo_pagamento,
                'status': 'NÃO ENCONTRADA',
                'valor_gerado': None,
                'diferenca': None,
                'os_correspondente': None,
                'observacao': f'Data {data_cartao} não encontrada nos dados gerados'
            })
            continue
        
        # Procura por valor na coluna correspondente
        campo_procurado = 'CARTÃO' if tipo_pagamento == 'CARTÃO' else 'PIX'
        valor_encontrado = None
        os_correspondente = None
        
        for _, gen_row in matching_generated.iterrows():
            if campo_procurado in gen_row.index:
                valor_gen = gen_row[campo_procurado]
                if pd.notna(valor_gen) and abs(valor_gen - valor_cartao) <= (valor_cartao * 0.01):  # 1% tolerância
                    valor_encontrado = valor_gen
                    os_correspondente = gen_row.get('N° OS', 'N/A')
                    break
        
        if valor_encontrado is not None:
            # Valor encontrado
            diferenca = abs(valor_cartao - valor_encontrado)
            is_match = diferenca <= (valor_cartao * 0.01)
            
            results.append({
                'identificador': identificador,
                'data_cartao': data_cartao,
                'valor_cartao': valor_cartao,
                'tipo_pagamento': tipo_pagamento,
                'status': 'COINCIDENTE' if is_match else 'DIVERGENTE',
                'valor_gerado': valor_encontrado,
                'diferenca': diferenca,
                'os_correspondente': os_correspondente,
                'observacao': f'Encontrado na coluna {campo_procurado}'
            })
        else:
            # Valor não encontrado
            results.append({
                'identificador': identificador,
                'data_cartao': data_cartao,
                'valor_cartao': valor_cartao,
                'tipo_pagamento': tipo_pagamento,
                'status': 'VALOR NÃO ENCONTRADO',
                'valor_gerado': None,
                'diferenca': None,
                'os_correspondente': None,
                'observacao': f'Valor {valor_cartao} não encontrado na coluna {campo_procurado} para a data {data_cartao}'
            })
    
    return results


def normalize_remetente(rem):
    if not rem:
        return ''
    rem = rem.lower().strip()
    rem = ''.join(c for c in unicodedata.normalize('NFD', rem) if unicodedata.category(c) != 'Mn')
    return rem

def audit_pix_transactions(banco_transactions: List[PixTransaction], recebimentos_transactions: List[PixTransaction], tolerancia_valor=0.01, tolerancia_dias=2) -> List[Dict]:
    """
    Auditoria PIX: para cada transação do banco, procura nos recebimentos uma transação com:
    - Data: tolerância de 2 dias (configurável)
    - Valor: tolerância de centavos (configurável)
    - Se não encontrar correspondência individual, agrupa transações do mesmo remetente na mesma data e tenta com valor somado
    - Para valores exatamente iguais, tenta com tolerância maior de datas (15 dias) se não encontrar com 2 dias
    Cada transação só pode ser usada uma vez.
    
    Args:
        banco_transactions: Lista de transações do banco
        recebimentos_transactions: Lista de transações dos recebimentos
        tolerancia_valor: Tolerância para valores (padrão: 0.01 centavos)
        tolerancia_dias: Tolerância para datas em dias (padrão: 2 dias)
    """
    logger = logging.getLogger(__name__)
    results = []
    usados_receb = set()
    usados_banco = set()
    
    # Primeira passada: tenta correspondências individuais
    for idx_banco, banco_tx in enumerate(banco_transactions):
        if idx_banco in usados_banco:
            continue
            
        match_idx = None
        tolerancia_dias_usada = tolerancia_dias
        
        # Primeira tentativa: tolerância normal (2 dias)
        for idx_rec, rec_tx in enumerate(recebimentos_transactions):
            if idx_rec in usados_receb:
                continue
            
            # Compara data com tolerância de dias
            try:
                # Converte datas para datetime para comparação
                data_banco = pd.to_datetime(banco_tx.data, format='%d/%m/%Y')
                data_receb = pd.to_datetime(rec_tx.data, format='%d/%m/%Y')
                
                # Calcula diferença em dias
                diff_dias = abs((data_banco - data_receb).days)
                
                # Verifica se está dentro da tolerância de dias
                if diff_dias > tolerancia_dias:
                    continue
            except:
                # Se falhar a conversão de data, mantém comparação exata
                if banco_tx.data != rec_tx.data:
                    continue
            
            # Compara valor (com tolerância de centavos)
            if abs(banco_tx.valor - rec_tx.valor) <= tolerancia_valor:
                match_idx = idx_rec
                break
        
        # Segunda tentativa: se não encontrou, tenta com tolerância maior (15 dias) para valores exatamente iguais
        if match_idx is None:
            tolerancia_dias_usada = 15
            for idx_rec, rec_tx in enumerate(recebimentos_transactions):
                if idx_rec in usados_receb:
                    continue
                
                # Compara data com tolerância maior
                try:
                    data_banco = pd.to_datetime(banco_tx.data, format='%d/%m/%Y')
                    data_receb = pd.to_datetime(rec_tx.data, format='%d/%m/%Y')
                    diff_dias = abs((data_banco - data_receb).days)
                    
                    if diff_dias > tolerancia_dias_usada:
                        continue
                except:
                    if banco_tx.data != rec_tx.data:
                        continue
                
                # Compara valor (deve ser exatamente igual)
                if banco_tx.valor == rec_tx.valor:
                    match_idx = idx_rec
                    break
        
        if match_idx is not None:
            usados_receb.add(match_idx)
            usados_banco.add(idx_banco)
            rec_tx = recebimentos_transactions[match_idx]
            
            # Calcula diferença de dias para a observação
            try:
                data_banco_dt = pd.to_datetime(banco_tx.data, format='%d/%m/%Y')
                data_receb_dt = pd.to_datetime(rec_tx.data, format='%d/%m/%Y')
                diff_dias = abs((data_banco_dt - data_receb_dt).days)
                obs_dias = f" (diferença de {diff_dias} dia{'s' if diff_dias != 1 else ''})" if diff_dias > 0 else ""
                data_banco_normalizada = data_banco_dt.strftime('%d/%m/%Y')
            except:
                obs_dias = ""
                data_banco_normalizada = banco_tx.data
            
            # Determina o status baseado na tolerância usada
            if tolerancia_dias_usada > tolerancia_dias:
                status = 'CORRESPONDÊNCIA ENCONTRADA (TOLERÂNCIA ESTENDIDA)'
                obs_tolerancia = f" - tolerância estendida para {tolerancia_dias_usada} dias"
            else:
                status = 'CORRESPONDÊNCIA ENCONTRADA'
                obs_tolerancia = ""
            
            results.append({
                'data_banco': data_banco_normalizada,
                'valor_banco': banco_tx.valor,
                'remetente_banco': banco_tx.remetente,
                'detalhes_banco': banco_tx.descricao,
                'data_recebimentos': rec_tx.data,
                'valor_recebimentos': rec_tx.valor,
                'detalhes_recebimentos': f'R$ {rec_tx.valor:,.2f} - OS: {getattr(rec_tx, "referencia", "N/A")}',
                'os_recebimentos': getattr(rec_tx, 'referencia', 'N/A'),
                'status': status,
                'observacao': f'Valor R$ {banco_tx.valor:,.2f} encontrado em recebimento OS: {getattr(rec_tx, "referencia", "N/A")}{obs_dias}{obs_tolerancia}'
            })
    
    # Segunda passada: agrupa transações não usadas do mesmo remetente na mesma data
    # Agrupa transações do banco por remetente e data
    grupos_banco = {}
    for idx_banco, banco_tx in enumerate(banco_transactions):
        if idx_banco in usados_banco:
            continue
            
        chave = (banco_tx.remetente, banco_tx.data)
        if chave not in grupos_banco:
            grupos_banco[chave] = []
        grupos_banco[chave].append((idx_banco, banco_tx))
    
    # Para cada grupo, tenta encontrar correspondência com valor somado
    for (remetente, data), transacoes in grupos_banco.items():
        if len(transacoes) == 1:
            # Transação única - não marca como "SEM CORRESPONDÊNCIA" aqui, deixa para a terceira passada tentar
            continue
        else:
            # Múltiplas transações - soma os valores e procura correspondência
            valor_total = sum(tx.valor for _, tx in transacoes)
            indices_banco = [idx for idx, _ in transacoes]
            
            # Procura correspondência com valor somado
            match_idx = None
            for idx_rec, rec_tx in enumerate(recebimentos_transactions):
                if idx_rec in usados_receb:
                    continue
                
                # Compara data com tolerância de dias
                try:
                    data_banco = pd.to_datetime(data, format='%d/%m/%Y')
                    data_receb = pd.to_datetime(rec_tx.data, format='%d/%m/%Y')
                    diff_dias = abs((data_banco - data_receb).days)
                    if diff_dias > tolerancia_dias:
                        continue
                except:
                    if data != rec_tx.data:
                        continue
                
                # Compara valor somado
                if abs(valor_total - rec_tx.valor) <= tolerancia_valor:
                    match_idx = idx_rec
                    break
            
            if match_idx is not None:
                # Correspondência encontrada com valor somado
                usados_receb.add(match_idx)
                usados_banco.update(indices_banco)
                rec_tx = recebimentos_transactions[match_idx]
                
                # Calcula diferença de dias
                try:
                    data_banco_dt = pd.to_datetime(data, format='%d/%m/%Y')
                    data_receb_dt = pd.to_datetime(rec_tx.data, format='%d/%m/%Y')
                    diff_dias = abs((data_banco_dt - data_receb_dt).days)
                    obs_dias = f" (diferença de {diff_dias} dia{'s' if diff_dias != 1 else ''})" if diff_dias > 0 else ""
                    data_banco_normalizada = data_banco_dt.strftime('%d/%m/%Y')
                except:
                    obs_dias = ""
                    data_banco_normalizada = data
                
                # Cria detalhes das transações agrupadas
                detalhes_transacoes = []
                for _, tx in transacoes:
                    detalhes_transacoes.append(f"R$ {tx.valor:,.2f}")
                detalhes_str = " + ".join(detalhes_transacoes)
                
                results.append({
                    'data_banco': data_banco_normalizada,
                    'valor_banco': valor_total,
                    'remetente_banco': remetente,
                    'detalhes_banco': f"Múltiplas transações: {detalhes_str}",
                    'data_recebimentos': rec_tx.data,
                    'valor_recebimentos': rec_tx.valor,
                    'detalhes_recebimentos': f'R$ {rec_tx.valor:,.2f} - OS: {getattr(rec_tx, "referencia", "N/A")}',
                    'os_recebimentos': getattr(rec_tx, 'referencia', 'N/A'),
                    'status': 'CORRESPONDÊNCIA ENCONTRADA (VALOR SOMADO)',
                    'observacao': f'Valor total R$ {valor_total:,.2f} ({detalhes_str}) encontrado em recebimento OS: {getattr(rec_tx, "referencia", "N/A")}{obs_dias}'
                })
            else:
                # Não encontrou correspondência - marca cada transação individualmente
                for idx_banco, banco_tx in transacoes:
                    results.append({
                        'data_banco': banco_tx.data,
                        'valor_banco': banco_tx.valor,
                        'remetente_banco': banco_tx.remetente,
                        'detalhes_banco': banco_tx.descricao,
                        'data_recebimentos': None,
                        'valor_recebimentos': None,
                        'detalhes_recebimentos': None,
                        'os_recebimentos': None,
                        'status': 'SEM CORRESPONDÊNCIA',
                        'observacao': f'Sem correspondência individual ou somada PIX de {remetente}'
                    })
    
    # Terceira passada: agrupa recebimentos não usados do mesmo dia e tenta encontrar correspondência com valor somado
    recebimentos_nao_usados = []
    for idx_rec, rec_tx in enumerate(recebimentos_transactions):
        if idx_rec not in usados_receb:
            recebimentos_nao_usados.append((idx_rec, rec_tx))
    
    # Agrupa recebimentos não usados por data
    grupos_recebimentos = {}
    for idx_rec, rec_tx in recebimentos_nao_usados:
        data = rec_tx.data
        if data not in grupos_recebimentos:
            grupos_recebimentos[data] = []
        grupos_recebimentos[data].append((idx_rec, rec_tx))
    
    # Para cada grupo de recebimentos, tenta encontrar correspondência com valor somado
    for data, transacoes in grupos_recebimentos.items():
        if len(transacoes) == 1:
            # Recebimento único - marca como não encontrado
            idx_rec, rec_tx = transacoes[0]
            results.append({
                'data_banco': None,
                'valor_banco': None,
                'remetente_banco': None,
                'detalhes_banco': None,
                'data_recebimentos': rec_tx.data,
                'valor_recebimentos': rec_tx.valor,
                'detalhes_recebimentos': f'R$ {rec_tx.valor:,.2f} - OS: {getattr(rec_tx, "referencia", "N/A")}',
                'os_recebimentos': getattr(rec_tx, 'referencia', 'N/A'),
                'status': 'SEM CORRESPONDÊNCIA',
                'observacao': f'Recebimento sem correspondência no banco'
            })
        else:
            # Múltiplos recebimentos - soma os valores e procura correspondência
            valor_total = sum(tx.valor for _, tx in transacoes)
            indices_receb = [idx for idx, _ in transacoes]
            # Procura correspondência com valor somado (usando tolerância baixa)
            match_idx = None
            for idx_banco, banco_tx in enumerate(banco_transactions):
                if idx_banco in usados_banco:
                    continue
                # Compara data com tolerância baixa
                try:
                    data_banco = pd.to_datetime(banco_tx.data, format='%d/%m/%Y')
                    data_receb = pd.to_datetime(data, format='%d/%m/%Y')
                    diff_dias = abs((data_banco - data_receb).days)
                    if diff_dias > tolerancia_dias:  # Usa tolerância baixa (2 dias)
                        continue
                except:
                    if banco_tx.data != data:
                        continue
                # Compara valor somado
                if abs(valor_total - banco_tx.valor) <= tolerancia_valor:
                    match_idx = idx_banco
                    break
            if match_idx is not None:
                # Correspondência encontrada com valor somado
                usados_banco.add(match_idx)
                usados_receb.update(indices_receb)
                banco_tx = banco_transactions[match_idx]
                # Calcula diferença de dias
                try:
                    data_banco_dt = pd.to_datetime(banco_tx.data, format='%d/%m/%Y')
                    data_receb_dt = pd.to_datetime(data, format='%d/%m/%Y')
                    diff_dias = abs((data_banco_dt - data_receb_dt).days)
                    obs_dias = f" (diferença de {diff_dias} dia{'s' if diff_dias != 1 else ''})" if diff_dias > 0 else ""
                    data_banco_normalizada = data_banco_dt.strftime('%d/%m/%Y')
                except:
                    obs_dias = ""
                    data_banco_normalizada = banco_tx.data
                # Cria detalhes dos recebimentos agrupados
                detalhes_recebimentos = []
                for _, tx in transacoes:
                    detalhes_recebimentos.append(f"OS: {getattr(tx, 'referencia', 'N/A')} (R$ {tx.valor:,.2f})")
                detalhes_str = " + ".join(detalhes_recebimentos)
                results.append({
                    'data_banco': data_banco_normalizada,
                    'valor_banco': banco_tx.valor,
                    'remetente_banco': banco_tx.remetente,
                    'detalhes_banco': banco_tx.descricao,
                    'data_recebimentos': data,
                    'valor_recebimentos': valor_total,
                    'detalhes_recebimentos': f'R$ {valor_total:,.2f} - {detalhes_str}',
                    'os_recebimentos': " + ".join([getattr(tx, 'referencia', 'N/A') for _, tx in transacoes]),
                    'status': 'CORRESPONDÊNCIA ENCONTRADA (RECEBIMENTOS SOMADOS)',
                    'observacao': f'Valor total R$ {valor_total:,.2f} ({detalhes_str}) encontrado em transação PIX de {banco_tx.remetente}{obs_dias}'
                })
            else:
                # Não encontrou correspondência - marca cada recebimento individualmente
                for idx_rec, rec_tx in transacoes:
                    results.append({
                        'data_banco': None,
                        'valor_banco': None,
                        'remetente_banco': None,
                        'detalhes_banco': None,
                        'data_recebimentos': rec_tx.data,
                        'valor_recebimentos': rec_tx.valor,
                        'detalhes_recebimentos': f'R$ {rec_tx.valor:,.2f} - OS: {getattr(rec_tx, "referencia", "N/A")}',
                        'os_recebimentos': getattr(rec_tx, 'referencia', 'N/A'),
                        'status': 'SEM CORRESPONDÊNCIA',
                        'observacao': f'Recebimento sem correspondência individual ou somada no banco'
                    })
    
    # Quarta passada: marca transações do banco não usadas como "SEM CORRESPONDÊNCIA"
    for idx_banco, banco_tx in enumerate(banco_transactions):
        if idx_banco not in usados_banco:
            results.append({
                'data_banco': banco_tx.data,
                'valor_banco': banco_tx.valor,
                'remetente_banco': banco_tx.remetente,
                'detalhes_banco': banco_tx.descricao,
                'data_recebimentos': None,
                'valor_recebimentos': None,
                'detalhes_recebimentos': None,
                'os_recebimentos': None,
                'status': 'SEM CORRESPONDÊNCIA',
                'observacao': f'Sem correspondência PIX de {banco_tx.remetente}'
            })
    
    return results


def group_pix_transactions_by_remetente(transactions: List[PixTransaction]) -> List[GroupedPixTransaction]:
    """Agrupa transações PIX da mesma pessoa no mesmo dia"""
    logger = logging.getLogger(__name__)
    
    # Agrupa por data (simplificado)
    grouped_dict = {}
    
    for tx in transactions:
        # Cria chave de agrupamento: apenas data
        group_key = tx.data
        
        if group_key not in grouped_dict:
            grouped_dict[group_key] = []
        grouped_dict[group_key].append(tx)
    
    # Cria transações agrupadas
    grouped_transactions = []
    
    for data, transacoes in grouped_dict.items():
        if len(transacoes) == 1:
            # Transação única - mantém como está
            tx = transacoes[0]
            grouped_tx = GroupedPixTransaction(
                data=data,
                valor_total=tx.valor,
                remetente=tx.remetente or "Desconhecido",
                origem=tx.origem,
                transacoes_originais=transacoes,
                quantidade_transacoes=1,
                referencia=tx.referencia
            )
        else:
            # Múltiplas transações no mesmo dia - agrupa
            valor_total = sum(tx.valor for tx in transacoes)
            # Tenta identificar um remetente comum ou usa "Múltiplos"
            remetentes = [tx.remetente for tx in transacoes if tx.remetente]
            if len(set(remetentes)) == 1 and remetentes[0]:
                remetente = remetentes[0]
            else:
                remetente = "Múltiplos remetentes"
            
            grouped_tx = GroupedPixTransaction(
                data=data,
                valor_total=valor_total,
                remetente=remetente,
                origem=transacoes[0].origem,
                transacoes_originais=transacoes,
                quantidade_transacoes=len(transacoes),
                referencia="Múltiplas transações"
            )
            logger.info(f"Agrupadas {len(transacoes)} transações em {data} - Total: R$ {valor_total:,.2f}")
        
        grouped_transactions.append(grouped_tx)
    
    logger.info(f"Transações agrupadas: {len(transactions)} -> {len(grouped_transactions)} grupos")
    return grouped_transactions


def group_recebimentos_by_os(transactions: List[PixTransaction]) -> List[GroupedPixTransaction]:
    """Agrupa transações de recebimentos por OS (referencia)"""
    logger = logging.getLogger(__name__)
    grouped_dict = {}
    for tx in transactions:
        group_key = tx.referencia or 'N/A'
        if group_key not in grouped_dict:
            grouped_dict[group_key] = []
        grouped_dict[group_key].append(tx)
    grouped_transactions = []
    for os_num, transacoes in grouped_dict.items():
        if len(transacoes) == 1:
            tx = transacoes[0]
            grouped_tx = GroupedPixTransaction(
                data=tx.data,
                valor_total=tx.valor,
                remetente="Recebimento",
                origem=tx.origem,
                transacoes_originais=transacoes,
                quantidade_transacoes=1,
                referencia=os_num
            )
        else:
            valor_total = sum(tx.valor for tx in transacoes)
            grouped_tx = GroupedPixTransaction(
                data=transacoes[0].data,
                valor_total=valor_total,
                remetente="Recebimentos múltiplos",
                origem=transacoes[0].origem,
                transacoes_originais=transacoes,
                quantidade_transacoes=len(transacoes),
                referencia=os_num
            )
        grouped_transactions.append(grouped_tx)
    logger.info(f"Transações de recebimentos agrupadas por OS: {len(transactions)} -> {len(grouped_transactions)} grupos")
    return grouped_transactions


def generate_unified_report_json(cartao_results, pix_results, cartao_stats, recebimentos_transactions, banco_transactions, output_file, banco_pix_csv, nfse_df=None, nfse_results=None):
    """
    Gera relatório unificado em JSON com resumo e detalhes das auditorias.
    """
    try:
        # Resumo executivo
        resumo = {
            'cartao': cartao_stats,
            'pix': {
                'total_transacoes_banco': len(banco_transactions),
                'total_transacoes_recebimentos': len(recebimentos_transactions),
                'correspondencias_encontradas': len([r for r in pix_results if 'CORRESPONDÊNCIA ENCONTRADA' in r['status']]),
                'sem_correspondencia': len([r for r in pix_results if r['status'] == 'SEM CORRESPONDÊNCIA']),
            },
            'nfse': {
                'total_notas': len(nfse_df) if nfse_df is not None else 0,
                'total_correspondencias': len([r for r in nfse_results if r['status'] == 'COINCIDENTE']) if nfse_results is not None else 0,
                'nao_encontradas': len([r for r in nfse_results if r['status'] == 'NÃO ENCONTRADA']) if nfse_results is not None else 0,
            } if nfse_df is not None and nfse_results is not None else {},
        }
        relatorio = {
            'resumo': resumo,
            'detalhes_cartao': cartao_results,
            'detalhes_pix': pix_results,
            'detalhes_nfse': nfse_results,
        }
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(relatorio, f, ensure_ascii=False, indent=2, default=str)
        print(f"[OK] Relatório unificado gerado em: {output_file}")
    except Exception as e:
        print(f"[ERRO] Erro ao gerar relatório JSON: {e}")
        raise


def configure_worksheet_properties(worksheet, sheet_name):
    """Configura propriedades da planilha para melhor apresentação"""
    from openpyxl.worksheet.views import SheetView
    
    # Configura view da planilha usando a API correta
    if not hasattr(worksheet, 'sheet_view') or worksheet.sheet_view is None:
        worksheet.sheet_view = SheetView()
    
    # Configura propriedades da view
    worksheet.sheet_view.showGridLines = True
    worksheet.sheet_view.showRowColHeaders = True
    worksheet.sheet_view.zoomScale = 100
    worksheet.sheet_view.zoomScaleNormal = 100
    worksheet.sheet_view.zoomScalePageLayoutView = 100
    
    # Configura propriedades específicas por tipo de planilha
    if 'Detalhes' in sheet_name:
        # Para detalhes, ajusta zoom para melhor visualização
        worksheet.sheet_view.zoomScale = 90
    elif 'Divergências' in sheet_name or 'Sem Correspondência' in sheet_name:
        # Para divergências, zoom menor para ver mais dados
        worksheet.sheet_view.zoomScale = 85


def safe_to_excel(df, writer, sheet_name, theme, border_config):
    """Salva DataFrame no Excel com formatação segura e otimizada"""
    # Processa o DataFrame para evitar problemas
    df_processed = df.copy()
    
    # Preenche valores NaN
    df_processed = df_processed.fillna('')
    
    # Converte para string e trata valores que começam com "="
    for col in df_processed.columns:
        df_processed[col] = df_processed[col].astype(str).apply(
            lambda x: "'" + x if isinstance(x, str) and x.startswith('=') else x
        )
    
    # Remove linhas completamente vazias
    df_processed = df_processed.dropna(how='all')
    
    # Se o DataFrame ficou vazio, cria uma linha com mensagem
    if df_processed.empty:
        df_processed = pd.DataFrame({'Mensagem': ['Nenhum dado disponível para esta seção']})
    
    # Salva no Excel
    df_processed.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Obtém a planilha e aplica formatação
    worksheet = writer.sheets[sheet_name]
    apply_worksheet_formatting(worksheet, df_processed, theme, border_config)
    configure_worksheet_properties(worksheet, sheet_name)


def optimize_column_widths(worksheet, df):
    """Otimiza a largura das colunas baseada no conteúdo e configurações"""
    from openpyxl.utils import get_column_letter
    
    for col in range(1, len(df.columns) + 1):
        column_name = df.columns[col - 1]
        
        # Largura mínima baseada na configuração
        min_width = COLUMN_WIDTHS.get(column_name, COLUMN_WIDTHS['default'])
        
        # Calcula largura baseada no conteúdo
        header_length = len(str(column_name))
        max_content_length = header_length
        
        # Analisa o conteúdo das células
        for row in range(min(len(df), 100)):  # Limita a 100 linhas para performance
            try:
                cell_value = str(df.iloc[row, col - 1])
                # Remove caracteres especiais para cálculo mais preciso
                clean_value = cell_value.replace('R$', '').replace(',', '').replace('.', '')
                max_content_length = max(max_content_length, len(clean_value))
            except:
                continue
        
        # Aplica fatores de ajuste baseados no tipo de coluna
        if any(keyword in column_name.lower() for keyword in ['observacao', 'descricao', 'notes']):
            # Colunas de texto longo - largura maior
            content_width = max_content_length * 1.3
            max_width = 100
        elif any(keyword in column_name.lower() for keyword in ['valor', 'diferenca', 'percentual']):
            # Colunas numéricas - largura fixa para formatação
            content_width = max_content_length * 1.1
            max_width = 25
        elif 'data' in column_name.lower():
            # Colunas de data - largura fixa
            content_width = 15
            max_width = 20
        else:
            # Colunas padrão
            content_width = max_content_length * 1.2
            max_width = 50
        
        # Define largura final
        final_width = max(min_width, min(content_width, max_width))
        worksheet.column_dimensions[get_column_letter(col)].width = final_width


def apply_worksheet_formatting(worksheet, df, theme, border_config):
    """Aplica formatação uniforme à planilha com largura de colunas otimizada"""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    # Configurações de borda
    border_style = Side(style=border_config['data_border'], color=border_config['border_color'])
    header_border_style = Side(style=border_config['header_border'], color=border_config['border_color'])
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color=theme['header_font'])
    header_fill = PatternFill(start_color=theme['header_bg'], end_color=theme['header_bg'], fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Estilo das células de dados
    data_font = Font(color='000000')
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    # Aplica formatação ao cabeçalho
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(
            left=header_border_style, right=header_border_style,
            top=header_border_style, bottom=header_border_style
        )
    
    # Aplica formatação aos dados
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = Border(
                left=border_style, right=border_style,
                top=border_style, bottom=border_style
            )
            
            # Formatação específica para colunas numéricas
            column_name = df.columns[col - 1]
            if any(keyword in column_name.lower() for keyword in ['valor', 'diferenca', 'percentual']):
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = CURRENCY_FORMATS['BRL']
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Formatação para datas
            elif 'data' in column_name.lower():
                if cell.value is not None:
                    cell.number_format = DATE_FORMATS['pt_BR']
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Otimiza largura das colunas
    optimize_column_widths(worksheet, df)


def audit_nfse_vs_recebimentos(nfse_df: pd.DataFrame, recebimentos_path: str) -> List[Dict]:
    """
    Audita a comparação entre notas fiscais e valores de mão de obra dos recebimentos
    
    Args:
        nfse_df: DataFrame com dados das notas fiscais
        recebimentos_df: DataFrame com dados dos recebimentos
        
    Returns:
        List[Dict]: Lista com resultados da auditoria
    """
    logger = logging.getLogger(__name__)
    logger.info("Iniciando auditoria NFSe vs Recebimentos...")
    
    results = []
    
    try:
        # Carrega dados de recebimentos baseado na extensão do arquivo
        if recebimentos_path.lower().endswith('.json'):
            recebimentos_df = pd.read_json(recebimentos_path, orient='records')
            logger.info(f"Carregados {len(recebimentos_df)} registros de recebimentos do JSON")
        else:
            recebimentos_df = pd.read_excel(recebimentos_path)
            logger.info(f"Carregados {len(recebimentos_df)} registros de recebimentos do Excel")
        
        # Normaliza colunas dos recebimentos
        recebimentos_df = recebimentos_df.copy()
        
        # Converte DATA PGTO para string para comparação
        if 'DATA PGTO' in recebimentos_df.columns:
            recebimentos_df['DATA_PGTO_STR'] = pd.to_datetime(recebimentos_df['DATA PGTO']).dt.strftime('%d/%m/%Y')
        
        # Calcula valor líquido (MÃO DE OBRA + DESCONTO) para cada recebimento
        if 'VALOR MÃO DE OBRA' in recebimentos_df.columns and 'DESCONTO' in recebimentos_df.columns:
            recebimentos_df['VALOR_LIQUIDO'] = recebimentos_df['VALOR MÃO DE OBRA'] + recebimentos_df['DESCONTO']
        else:
            recebimentos_df['VALOR_LIQUIDO'] = recebimentos_df.get('VALOR MÃO DE OBRA', 0)
        
        # Processa cada nota fiscal
        for _, nfse_row in nfse_df.iterrows():
            numero_nfse = nfse_row.get('numero_nfse')
            nome_tomador = nfse_row.get('nome_tomador')
            valor_nfse = nfse_row.get('valor_total')
            data_nfse = nfse_row.get('data_emissao')
            
            if not all([numero_nfse, valor_nfse, data_nfse]):
                continue
            
            # Converte valor da NFSe para float (já deve estar como float, mas garante)
            try:
                if isinstance(valor_nfse, (int, float)):
                    valor_nfse_float = float(valor_nfse)
                else:
                    valor_nfse_float = float(str(valor_nfse).replace('R$', '').replace('.', '').replace(',', '.').strip())
            except (ValueError, AttributeError):
                valor_nfse_float = 0
            
            # Procura correspondência nos recebimentos
            matching_recebimentos = []
            
            # Busca por valor líquido (comparação exata com round(2)) - sem restrição de data
            if 'VALOR_LIQUIDO' in recebimentos_df.columns:
                matching_recebimentos = recebimentos_df[
                    recebimentos_df['VALOR_LIQUIDO'].round(2) == round(valor_nfse_float, 2)
                ]
            
            # Determina status da auditoria
            if len(matching_recebimentos) == 1:
                # Correspondência exata encontrada
                recebimento = matching_recebimentos.iloc[0]
                valor_recebimento = recebimento['VALOR_LIQUIDO'] if 'VALOR_LIQUIDO' in recebimento.index else 0
                mao_obra = recebimento['VALOR MÃO DE OBRA'] if 'VALOR MÃO DE OBRA' in recebimento.index else 0
                desconto = recebimento['DESCONTO'] if 'DESCONTO' in recebimento.index else 0
                data_recebimento = recebimento['DATA_PGTO_STR'] if 'DATA_PGTO_STR' in recebimento.index else 'N/A'
                diferenca = valor_nfse_float - valor_recebimento
                status = 'COINCIDENTE'
                os_correspondente = recebimento['N° OS'] if 'N° OS' in recebimento.index else 'N/A'
                
            elif len(matching_recebimentos) > 1:
                # Múltiplas correspondências
                recebimento = matching_recebimentos.iloc[0]  # Pega o primeiro
                valor_recebimento = recebimento['VALOR_LIQUIDO'] if 'VALOR_LIQUIDO' in recebimento.index else 0
                mao_obra = recebimento['VALOR MÃO DE OBRA'] if 'VALOR MÃO DE OBRA' in recebimento.index else 0
                desconto = recebimento['DESCONTO'] if 'DESCONTO' in recebimento.index else 0
                data_recebimento = recebimento['DATA_PGTO_STR'] if 'DATA_PGTO_STR' in recebimento.index else 'N/A'
                diferenca = valor_nfse_float - valor_recebimento
                status = 'MÚLTIPLAS CORRESPONDÊNCIAS'
                os_list = []
                for r in matching_recebimentos.head(3).itertuples():
                    os_list.append(str(getattr(r, 'N° OS', 'N/A')))
                os_correspondente = f"Múltiplas OS: {', '.join(os_list)}"
                
            else:
                # Nenhuma correspondência encontrada
                valor_recebimento = 0
                mao_obra = 0
                desconto = 0
                data_recebimento = 'N/A'
                diferenca = valor_nfse_float
                status = 'NÃO ENCONTRADA'
                os_correspondente = 'N/A'
            
            # Calcula diferença percentual
            dif_percentual = (diferenca / valor_nfse_float * 100) if valor_nfse_float > 0 else 0
            
            # Cria resultado da auditoria
            result = {
                'numero_nfse': numero_nfse,
                'nome_tomador': nome_tomador,
                'valor_nfse': valor_nfse_float,
                'data_nfse': data_nfse,
                'valor_mao_obra': mao_obra,
                'desconto': desconto,
                'valor_liquido': valor_recebimento,
                'data_recebimento': data_recebimento,
                'diferenca': diferenca,
                'dif_percentual': dif_percentual,
                'status': status,
                'os_correspondente': os_correspondente,
                'observacao': f"NFSe {numero_nfse} - {nome_tomador}"
            }
            
            results.append(result)
        
        logger.info(f"Auditoria NFSe vs Recebimentos concluída: {len(results)} registros processados")
        
        # Estatísticas
        coincidentes = len([r for r in results if r['status'] == 'COINCIDENTE'])
        nao_encontradas = len([r for r in results if r['status'] == 'NÃO ENCONTRADA'])
        multiplas = len([r for r in results if r['status'] == 'MÚLTIPLAS CORRESPONDÊNCIAS'])
        
        logger.info(f"  Coincidentes: {coincidentes}")
        logger.info(f"  Não encontradas: {nao_encontradas}")
        logger.info(f"  Múltiplas correspondências: {multiplas}")
        
        return results
        
    except Exception as e:
        logger.error(f"Erro na auditoria NFSe vs Recebimentos: {e}")
        return []


def extract_remetente_from_description(descricao: str) -> Optional[str]:
    """Extrai o nome do remetente da descrição da transação PIX"""
    try:
        # Padrões específicos baseados no formato real do CSV
        patterns = [
            # Padrão: "Transferência recebida pelo Pix - NOME - CPF/CNPJ - BANCO"
            r'Transferência recebida pelo Pix\s*-\s*([^-]+?)\s*-\s*[•\d\./-]+\s*-',
            # Padrão: "Transferência Recebida - NOME - CPF/CNPJ - BANCO"
            r'Transferência Recebida\s*-\s*([^-]+?)\s*-\s*[•\d\./-]+\s*-',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, descricao, re.IGNORECASE)
            if match:
                remetente = match.group(1).strip()
                # Remove caracteres especiais e normaliza
                remetente = re.sub(r'[^\w\s]', '', remetente).strip()
                # Remove espaços extras
                remetente = re.sub(r'\s+', ' ', remetente)
                if len(remetente) > 2:  # Nome deve ter pelo menos 3 caracteres
                    return remetente
        
        # Se não encontrou com os padrões específicos, tenta extrair o nome antes do primeiro CPF/CNPJ
        if '•••' in descricao or re.search(r'\d{3}\.\d{3}\.\d{3}', descricao):
            # Procura por texto antes do CPF/CNPJ
            parts = descricao.split(' - ')
            if len(parts) >= 2:
                # Pega a segunda parte (após "Transferência recebida pelo Pix")
                nome_part = parts[1]
                # Remove o CPF/CNPJ se presente
                nome_clean = re.sub(r'[•\d\./-]+', '', nome_part).strip()
                if len(nome_clean) > 2:
                    return nome_clean
        
        return None
    except:
        return None


def extract_chave_pix_from_description(descricao: str) -> Optional[str]:
    """Extrai a chave PIX da descrição da transação"""
    try:
        # Padrões para CPF, CNPJ, email, telefone
        patterns = [
            r'(\d{3}\.\d{3}\.\d{3}-\d{2})',  # CPF
            r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})',  # CNPJ
            r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})',  # Email
            r'(\+55\s?\d{2}\s?\d{4,5}\s?\d{4})',  # Telefone
        ]
        
        for pattern in patterns:
            match = re.search(pattern, descricao)
            if match:
                return match.group(1)
        
        return None
    except:
        return None


def executar_auditoria(cartao_csv: str, banco_csv: str, recebimentos_path: str, nfse_directory: str = None, output_file: str = None):
    """
    Executa a auditoria unificada com os arquivos especificados
    
    Args:
        cartao_csv: Caminho para o arquivo CSV de transações de cartão
        banco_csv: Caminho para o arquivo CSV de transações PIX do banco
        recebimentos_excel: Caminho para o arquivo Excel de recebimentos
        nfse_directory: Caminho para a pasta das notas fiscais (NFSe) - opcional
        output_file: Caminho para o arquivo de saída (opcional)
    """
    logger = setup_logging()
    
    try:
        logger.info("=== AUDITORIA UNIFICADA COMPLETA ===")
        
        # Define arquivo de saída padrão se não especificado
        if not output_file:
            output_file = "data/relatorios/auditoria_unificada_completa.xlsx"
        
        # Verifica se os arquivos existem
        if not os.path.exists(cartao_csv):
            raise FileNotFoundError(f"Arquivo CSV do cartão não encontrado: {cartao_csv}")
        
        if not os.path.exists(banco_csv):
            raise FileNotFoundError(f"Arquivo CSV do banco não encontrado: {banco_csv}")
        
        if not os.path.exists(recebimentos_path):
            raise FileNotFoundError(f"Arquivo de recebimentos não encontrado: {recebimentos_path}")
        
        # Verifica se a pasta das notas fiscais existe (se fornecida)
        if nfse_directory and not os.path.exists(nfse_directory):
            logger.warning(f"Pasta das notas fiscais não encontrada: {nfse_directory}")
            nfse_directory = None
        
        logger.info("Carregando dados...")
        logger.info(f"📄 Cartão: {os.path.basename(cartao_csv)}")
        logger.info(f"🏦 Banco: {os.path.basename(banco_csv)}")
        logger.info(f"📊 Recebimentos: {os.path.basename(recebimentos_path)}")
        if nfse_directory:
            logger.info(f"📋 Notas Fiscais: {os.path.basename(nfse_directory)}")
        
        # Carrega dados de cartão
        cartao_df = parse_cartao_csv(cartao_csv)
        
        # Carrega dados gerados baseado na extensão do arquivo
        auditor = DataAuditor(tolerance_percentage=0.01)
        if recebimentos_path.lower().endswith('.json'):
            generated_df = pd.read_json(recebimentos_path, orient='records')
            logger.info(f"Carregados {len(generated_df)} registros de recebimentos do JSON")
        else:
            generated_df = auditor.load_generated_data(recebimentos_path)
        generated_df = auditor.normalize_column_names(generated_df)
        
        # Converte DATA PGTO para date se necessário
        if 'DATA PGTO' in generated_df.columns:
            generated_df['DATA PGTO'] = pd.to_datetime(generated_df['DATA PGTO']).dt.date
        
        # Carrega dados PIX do JSON do banco (sempre deve existir)
        banco_json_dir = "data/json/banco"
        json_files = [f for f in os.listdir(banco_json_dir) if f.startswith('banco_') and f.endswith('.json')]
        if not json_files:
            raise Exception("Nenhum JSON do banco encontrado. Execute a conversão primeiro.")
        
        # Pega o arquivo mais recente
        json_files.sort(reverse=True)
        banco_json_path = os.path.join(banco_json_dir, json_files[0])
        logger.info(f"Usando JSON do banco: {banco_json_path}")
        banco_transactions = load_banco_pix_json(banco_json_path)
        
        if recebimentos_path.lower().endswith('.json'):
            recebimentos_transactions = load_recebimentos_json(recebimentos_path)
        else:
            recebimentos_transactions = load_recebimentos_excel(recebimentos_path)
        
        # Carrega dados das notas fiscais (se pasta fornecida)
        nfse_df = None
        if nfse_directory:
            try:
                logger.info("Carregando dados das notas fiscais...")
                from extrator_nfse import NFSeExtractor
                extrator = NFSeExtractor()
                nfse_df = extrator.process_directory(nfse_directory)
                logger.info(f"Notas fiscais carregadas: {len(nfse_df)} registros")
            except Exception as e:
                logger.warning(f"Erro ao carregar notas fiscais: {e}")
                nfse_df = None
        
        logger.info("Executando auditorias...")
        
        # Executa auditoria de cartão
        cartao_results = audit_cartao_transactions(cartao_df, generated_df)
        
        # Calcula estatísticas do cartão
        cartao_stats = {
            'total_transacoes': len(cartao_df),
            'cartao_encontradas': len([r for r in cartao_results if r['tipo_pagamento'] == 'CARTÃO' and r['status'] == 'COINCIDENTE']),
            'pix_encontradas': len([r for r in cartao_results if r['tipo_pagamento'] == 'PIX' and r['status'] == 'COINCIDENTE']),
            'nao_encontradas': len([r for r in cartao_results if r['status'] in ['NÃO ENCONTRADA', 'VALOR NÃO ENCONTRADO']]),
            'valores_coincidentes': len([r for r in cartao_results if r['status'] == 'COINCIDENTE']),
            'valores_divergentes': len([r for r in cartao_results if r['status'] == 'DIVERGENTE'])
        }
        
        # Executa auditoria PIX
        pix_results = audit_pix_transactions(banco_transactions, recebimentos_transactions)
        
        # Executa auditoria NFSe vs Recebimentos (se dados disponíveis)
        nfse_results = None
        if nfse_df is not None and not nfse_df.empty:
            logger.info("Executando auditoria NFSe vs Recebimentos...")
            nfse_results = audit_nfse_vs_recebimentos(nfse_df, recebimentos_path)
        
        logger.info("Gerando relatório unificado...")
        
        # Gera relatório unificado
        generate_unified_report_json(cartao_results, pix_results, cartao_stats, recebimentos_transactions, banco_transactions, output_file, banco_csv, nfse_df, nfse_results)
        
        logger.info(f"✅ Auditoria unificada concluída!")
        logger.info(f"📊 Relatório salvo em: {output_file}")
        
        # Exibe resumo no console
        logger.info("\n=== RESUMO EXECUTIVO ===")
        logger.info(f"Cartão - Total: {cartao_stats['total_transacoes']}, Coincidentes: {cartao_stats['valores_coincidentes']}")
        logger.info(f"PIX - Banco: {len(banco_transactions)}, Recebimentos: {len(recebimentos_transactions)}")
        logger.info(f"PIX - Correspondências: {len([r for r in pix_results if 'CORRESPONDÊNCIA ENCONTRADA' in r['status']])}")
        if nfse_df is not None:
            logger.info(f"NFSe - Total: {len(nfse_df)} notas fiscais")
        if nfse_results is not None and nfse_results:
            logger.info(f"NFSe vs Recebimentos - Coincidentes: {len([r for r in nfse_results if r['status'] == 'COINCIDENTE'])}")
            logger.info(f"NFSe vs Recebimentos - Não encontradas: {len([r for r in nfse_results if r['status'] == 'NÃO ENCONTRADA'])}")
        
        return output_file
        
    except Exception as e:
        logger.error(f"❌ Erro na auditoria: {e}")
        raise


def main():
    """Função principal"""
    logger = setup_logging()
    
    try:
        logger.info("=== AUDITORIA UNIFICADA COMPLETA ===")
        
        # Pergunta sobre o método de seleção de arquivos
        print("\n=== MÉTODO DE SELEÇÃO DE ARQUIVOS ===")
        print("1. Interface gráfica (recomendado)")
        print("2. PowerShell (linha de comando)")
        
        choice = input("\nEscolha o método (1 ou 2): ").strip()
        
        if choice == "1":
            files = select_files_gui()
        else:
            files = select_files_powershell()
        
        if not files:
            logger.info("Seleção de arquivos cancelada")
            return
        
        # Extrai caminhos dos arquivos
        cartao_csv = files['cartao_csv']
        banco_csv = files['banco_csv']
        recebimentos_path = files['recebimentos_excel']  # Mantém compatibilidade com GUI
        report_file = "data/relatorios/auditoria_unificada_completa.xlsx"
        
        # Verifica se os arquivos existem
        if not os.path.exists(cartao_csv):
            logger.error(f"Arquivo CSV do cartão não encontrado: {cartao_csv}")
            return
        
        if not os.path.exists(banco_csv):
            logger.error(f"Arquivo CSV do banco não encontrado: {banco_csv}")
            return
        
        if not os.path.exists(recebimentos_path):
            logger.error(f"Arquivo de recebimentos não encontrado: {recebimentos_path}")
            return
        
        logger.info("Carregando dados...")
        logger.info(f"📄 Cartão: {os.path.basename(cartao_csv)}")
        logger.info(f"🏦 Banco: {os.path.basename(banco_csv)}")
        logger.info(f"📊 Recebimentos: {os.path.basename(recebimentos_path)}")
        
        # Carrega dados de cartão
        cartao_df = parse_cartao_csv(cartao_csv)
        
        # Carrega dados gerados baseado na extensão do arquivo
        auditor = DataAuditor(tolerance_percentage=0.01)
        if recebimentos_path.lower().endswith('.json'):
            generated_df = pd.read_json(recebimentos_path, orient='records')
            logger.info(f"Carregados {len(generated_df)} registros de recebimentos do JSON")
        else:
            generated_df = auditor.load_generated_data(recebimentos_path)
        generated_df = auditor.normalize_column_names(generated_df)
        
        # Converte DATA PGTO para date se necessário
        if 'DATA PGTO' in generated_df.columns:
            generated_df['DATA PGTO'] = pd.to_datetime(generated_df['DATA PGTO']).dt.date
        
        # Carrega dados PIX do JSON do banco (sempre deve existir)
        banco_json_dir = "data/json/banco"
        json_files = [f for f in os.listdir(banco_json_dir) if f.startswith('banco_') and f.endswith('.json')]
        if not json_files:
            raise Exception("Nenhum JSON do banco encontrado. Execute a conversão primeiro.")
        
        # Pega o arquivo mais recente
        json_files.sort(reverse=True)
        banco_json_path = os.path.join(banco_json_dir, json_files[0])
        logger.info(f"Usando JSON do banco: {banco_json_path}")
        banco_transactions = load_banco_pix_json(banco_json_path)
        
        if recebimentos_path.lower().endswith('.json'):
            recebimentos_transactions = load_recebimentos_json(recebimentos_path)
        else:
            recebimentos_transactions = load_recebimentos_excel(recebimentos_path)
        
        logger.info("Executando auditorias...")
        
        # Executa auditoria de cartão
        cartao_results = audit_cartao_transactions(cartao_df, generated_df)
        
        # Calcula estatísticas do cartão
        cartao_stats = {
            'total_transacoes': len(cartao_df),
            'cartao_encontradas': len([r for r in cartao_results if r['tipo_pagamento'] == 'CARTÃO' and r['status'] == 'COINCIDENTE']),
            'pix_encontradas': len([r for r in cartao_results if r['tipo_pagamento'] == 'PIX' and r['status'] == 'COINCIDENTE']),
            'nao_encontradas': len([r for r in cartao_results if r['status'] in ['NÃO ENCONTRADA', 'VALOR NÃO ENCONTRADO']]),
            'valores_coincidentes': len([r for r in cartao_results if r['status'] == 'COINCIDENTE']),
            'valores_divergentes': len([r for r in cartao_results if r['status'] == 'DIVERGENTE'])
        }
        
        # Executa auditoria PIX
        pix_results = audit_pix_transactions(banco_transactions, recebimentos_transactions)
        
        # Executa auditoria NFSe vs Recebimentos (se dados disponíveis)
        nfse_results = None
        if nfse_df is not None and not nfse_df.empty:
            logger.info("Executando auditoria NFSe vs Recebimentos...")
            nfse_results = audit_nfse_vs_recebimentos(nfse_df, recebimentos_path)
        
        logger.info("Gerando relatório unificado...")
        
        # Gera relatório unificado
        generate_unified_report_json(cartao_results, pix_results, cartao_stats, recebimentos_transactions, banco_transactions, report_file, banco_csv, nfse_df, nfse_results)
        
        logger.info(f"✅ Auditoria unificada concluída!")
        logger.info(f"📊 Relatório salvo em: {report_file}")
        
        # Exibe resumo no console
        logger.info("\n=== RESUMO EXECUTIVO ===")
        logger.info(f"Cartão - Total: {cartao_stats['total_transacoes']}, Coincidentes: {cartao_stats['valores_coincidentes']}")
        logger.info(f"PIX - Banco: {len(banco_transactions)}, Recebimentos: {len(recebimentos_transactions)}")
        logger.info(f"PIX - Correspondências: {len([r for r in pix_results if 'CORRESPONDÊNCIA ENCONTRADA' in r['status']])}")
        if nfse_df is not None:
            logger.info(f"NFSe - Total: {len(nfse_df)} notas fiscais")
        if nfse_results is not None and nfse_results:
            logger.info(f"NFSe vs Recebimentos - Coincidentes: {len([r for r in nfse_results if r['status'] == 'COINCIDENTE'])}")
            logger.info(f"NFSe vs Recebimentos - Não encontradas: {len([r for r in nfse_results if r['status'] == 'NÃO ENCONTRADA'])}")
        
        # Mostra mensagem de sucesso na interface gráfica se disponível
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo("Sucesso", f"Auditoria concluída com sucesso!\n\nRelatório salvo em:\n{report_file}")
            root.destroy()
        except:
            pass
        
    except Exception as e:
        logger.error(f"❌ Erro na auditoria: {e}")
        
        # Mostra erro na interface gráfica se disponível
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Erro", f"Erro na auditoria:\n{e}")
            root.destroy()
        except:
            pass
        
        raise


if __name__ == '__main__':
    main() 