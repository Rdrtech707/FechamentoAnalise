import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from style_config import (
    CONTABEIS_COLS, CURRENCY_FORMATS, DATE_FORMATS, THEMES, 
    DECIMAL_SEPARATORS, COLUMN_WIDTHS, BORDER_STYLES, BORDER_CONFIGS
)


def export_to_excel(
    dataframes_by_month: dict,
    output_dir: str,
    currency: str = 'BRL',
    language: str = 'pt_BR',
    theme: str = 'default',
    decimal_separator: str = None,
    border_theme: str = 'default'
):
    """
    Salva cada DataFrame em planilhas Excel separadas por mês,
    ajustando automaticamente a largura das colunas e formatando
    colunas numéricas em estilo contábil com duas casas decimais.
    Permite customizar símbolo, separador decimal, tema de cores e bordas.
    """
    os.makedirs(output_dir, exist_ok=True)

    currency_format = CURRENCY_FORMATS.get(currency, 'R$ #,##0.00')
    date_format = DATE_FORMATS.get(language, 'dd/mm/yyyy')
    theme_cfg = THEMES.get(theme, THEMES['default'])
    border_cfg = BORDER_CONFIGS.get(border_theme, BORDER_CONFIGS['default'])
    decimal_sep = decimal_separator or DECIMAL_SEPARATORS.get(language, ',')

    for month, df in dataframes_by_month.items():
        filepath = os.path.join(output_dir, f"Recebimentos_{month}.xlsx")
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            sheet_name = month
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Estilo de cabeçalho
            header_fill = PatternFill(start_color=theme_cfg['header_bg'], end_color=theme_cfg['header_bg'], fill_type='solid')
            header_font = Font(color=theme_cfg['header_font'], bold=True)

            # Configurar bordas
            border_color = border_cfg['border_color']
            header_border_style = BORDER_STYLES.get(border_cfg['header_border'])
            data_border_style = BORDER_STYLES.get(border_cfg['data_border'])

            for idx, col in enumerate(df.columns, start=1):
                # Ajusta largura da coluna usando configuração personalizada
                column_width = COLUMN_WIDTHS.get(col, COLUMN_WIDTHS['default'])
                ws.column_dimensions[get_column_letter(idx)].width = column_width

                # Aplica formatação contábil para colunas numéricas
                if col in CONTABEIS_COLS:
                    for row_idx, cell in enumerate(ws[get_column_letter(idx)][1:], start=2):
                        cell.number_format = currency_format
                        cell.alignment = Alignment(horizontal='left')
                        cell.fill = PatternFill(start_color=theme_cfg['contabil_bg'], end_color=theme_cfg['contabil_bg'], fill_type='solid')
                        cell.font = Font(color=theme_cfg['contabil_font'])
                        
                        # Aplica bordas aos dados
                        if data_border_style:
                            cell.border = Border(
                                left=Side(style=data_border_style, color=border_color),
                                right=Side(style=data_border_style, color=border_color),
                                top=Side(style=data_border_style, color=border_color),
                                bottom=Side(style=data_border_style, color=border_color)
                            )

                # Aplica estilo ao cabeçalho
                header_cell = ws[f"{get_column_letter(idx)}1"]
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = Alignment(horizontal='center')
                
                # Aplica bordas ao cabeçalho
                if header_border_style:
                    header_cell.border = Border(
                        left=Side(style=header_border_style, color=border_color),
                        right=Side(style=header_border_style, color=border_color),
                        top=Side(style=header_border_style, color=border_color),
                        bottom=Side(style=header_border_style, color=border_color)
                    )

            # Ajusta separador decimal se necessário (apenas visual, não altera valores)
            # (Excel usa o separador do sistema, mas podemos ajustar o formato se necessário)
            # Não implementado aqui pois depende do Excel do usuário
